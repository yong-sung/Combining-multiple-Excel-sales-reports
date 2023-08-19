import glob
from openpyxl import load_workbook
from openpyxl import Workbook

판매보고들 = glob.glob(r"1.여러개의 엑셀 판매보고서 합치기\판매보고_*.xlsx") # 폴더 안에 있는 모든 판매보고_ 파일들을 리스트로 만들어줌
print(판매보고들)

# 데이터 저장을 위한 리스트 초기화
판매점_list = []
날짜_list = []
금액_list = []

# 엑셀 파일 처리 반복문
for 판매보고 in 판매보고들:
    wb = load_workbook(판매보고, data_only=True) # 엑셀 파일을 불러오고, 'active' 속성으로 현재 활성화된 시트를 가져옴.
    ws = wb.active
    판매점_list.append(ws['B1'].value) # 엑셀의 1행 2열 값을 리스트에 추가
    날짜_list.append(ws['B2'].value)  
    금액_list.append(ws['B3'].value)  
    
print(판매점_list)
print(날짜_list)
print(금액_list)

# 결과 엑셀 파일 처리
try:
    wb = load_workbook(r"1.여러개의 엑셀 판매보고서 합치기\결과.xlsx", data_only=True)
    ws = wb.active
except:
    wb = Workbook()
    ws = wb.active

# 데이터 저장
for i in range(len(판매점_list)):
    ws.cell(row=i+1,column=1).value = 판매점_list[i]
    ws.cell(row=i+1,column=2).value = 날짜_list[i]
    ws.cell(row=i+1,column=3).value = 금액_list[i]

# 결과 파일 저장
wb.save(r"1.여러개의 엑셀 판매보고서 합치기\결과.xlsx")