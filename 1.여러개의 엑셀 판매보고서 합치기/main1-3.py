import glob
from openpyxl import load_workbook
from openpyxl import Workbook

판매보고들 = glob.glob(r"1.여러개의 엑셀 판매보고서 합치기\판매보고_*.xlsx") # 폴더 안에 있는 모든 판매보고_ 파일들을 리스트로 만들어줌
print(판매보고들)

판매점_list = []
날짜_list = []
금액_list = []

for 판매보고 in 판매보고들:
    wb = load_workbook(판매보고, data_only=True)
    ws = wb.active
    판매점_list.append(ws['B1'].value)  
    날짜_list.append(ws['B2'].value)  
    금액_list.append(ws['B3'].value)  
    
print(판매점_list)
print(날짜_list)
print(금액_list)